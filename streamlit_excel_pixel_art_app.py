
import io
from dataclasses import dataclass
from typing import List, Optional, Tuple, Dict

import streamlit as st
from PIL import Image, ImageDraw
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

RGB = Tuple[int, int, int]
Pixel = Optional[RGB]

EXCEL_GREEN: RGB = (16, 124, 65)
WHITE: RGB = (255, 255, 255)
BLACK: RGB = (0, 0, 0)

def to_hex_argb(rgb: RGB) -> str:
    r, g, b = rgb
    return f"FF{r:02X}{g:02X}{b:02X}"

def clamp(x, lo=0, hi=255): 
    return max(lo, min(hi, x))

def lighten(rgb: RGB, amount: float) -> RGB:
    r, g, b = rgb
    return (clamp(int(r + (255 - r) * amount)),
            clamp(int(g + (255 - g) * amount)),
            clamp(int(b + (255 - b) * amount)))

def darken(rgb: RGB, amount: float) -> RGB:
    r, g, b = rgb
    return (clamp(int(r * (1 - amount))),
            clamp(int(g * (1 - amount))),
            clamp(int(b * (1 - amount))))

# ---------------- pixel font (5x7) for A-Z, 0-9 and a few symbols ----------------
# Each glyph is a list of 7 strings, each string 5 chars '1'/'0'
FONT_5x7 = {
    "A": ["01110","10001","10001","11111","10001","10001","10001"],
    "B": ["11110","10001","11110","10001","10001","10001","11110"],
    "C": ["01111","10000","10000","10000","10000","10000","01111"],
    "D": ["11110","10001","10001","10001","10001","10001","11110"],
    "E": ["11111","10000","11110","10000","10000","10000","11111"],
    "F": ["11111","10000","11110","10000","10000","10000","10000"],
    "G": ["01110","10000","10000","10111","10001","10001","01110"],
    "H": ["10001","10001","11111","10001","10001","10001","10001"],
    "I": ["11111","00100","00100","00100","00100","00100","11111"],
    "J": ["11111","00010","00010","00010","00010","10010","01100"],
    "K": ["10001","10010","11100","10100","10010","10001","10001"],
    "L": ["10000","10000","10000","10000","10000","10000","11111"],
    "M": ["10001","11011","10101","10101","10001","10001","10001"],
    "N": ["10001","11001","10101","10011","10001","10001","10001"],
    "O": ["01110","10001","10001","10001","10001","10001","01110"],
    "P": ["11110","10001","10001","11110","10000","10000","10000"],
    "Q": ["01110","10001","10001","10001","10101","10010","01101"],
    "R": ["11110","10001","10001","11110","10100","10010","10001"],
    "S": ["01111","10000","10000","01110","00001","00001","11110"],
    "T": ["11111","00100","00100","00100","00100","00100","00100"],
    "U": ["10001","10001","10001","10001","10001","10001","01110"],
    "V": ["10001","10001","10001","10001","10001","01010","00100"],
    "W": ["10001","10001","10001","10101","10101","11011","10001"],
    "X": ["10001","01010","00100","00100","00100","01010","10001"],
    "Y": ["10001","01010","00100","00100","00100","00100","00100"],
    "Z": ["11111","00001","00010","00100","01000","10000","11111"],
    "0": ["01110","10001","10011","10101","11001","10001","01110"],
    "1": ["00100","01100","00100","00100","00100","00100","01110"],
    "2": ["01110","10001","00001","00010","00100","01000","11111"],
    "3": ["11110","00001","00001","01110","00001","00001","11110"],
    "4": ["00010","00110","01010","10010","11111","00010","00010"],
    "5": ["11111","10000","11110","00001","00001","10001","01110"],
    "6": ["01110","10000","11110","10001","10001","10001","01110"],
    "7": ["11111","00001","00010","00100","01000","01000","01000"],
    "8": ["01110","10001","10001","01110","10001","10001","01110"],
    "9": ["01110","10001","10001","01111","00001","00001","01110"],
    " ": ["00000","00000","00000","00000","00000","00000","00000"],
    "-": ["00000","00000","00000","11111","00000","00000","00000"],
}

def normalize_ascii(s: str) -> str:
    repl = {
        "Ą":"A","Ć":"C","Ę":"E","Ł":"L","Ń":"N","Ó":"O","Ś":"S","Ż":"Z","Ź":"Z",
        "ą":"A","ć":"C","ę":"E","ł":"L","ń":"N","ó":"O","ś":"S","ż":"Z","ź":"Z",
        "—":"-","–":"-"
    }
    return "".join(repl.get(ch, ch).upper() if ch.isalpha() or ch in repl else ch for ch in s)

def text_size_cells(text: str, scale: int, spacing: int) -> int:
    w = 0
    for i, ch in enumerate(text):
        if ch in FONT_5x7:
            w += 5*scale
            if i < len(text)-1:
                w += spacing
    return w

def blit_glyph(grid, x, y, ch, color, scale=1):
    glyph = FONT_5x7.get(ch)
    if glyph is None:
        return x  # skip unknown
    rows, cols = 7, 5
    for gy in range(rows):
        row = glyph[gy]
        for gx in range(cols):
            if row[gx] == "1":
                for sy in range(scale):
                    for sx in range(scale):
                        yy = y + gy*scale + sy
                        xx = x + gx*scale + sx
                        if 0 <= yy < len(grid) and 0 <= xx < len(grid[0]):
                            grid[yy][xx] = color
    return x + cols*scale

def blit_text(grid, x, y, text, color, scale=1, spacing=1):
    for i, ch in enumerate(text):
        x = blit_glyph(grid, x, y, ch, color, scale)
        if i < len(text)-1:
            x += spacing
    return x

# ---------------- Mosaic from photo ----------------
@dataclass
class MosaicOptions:
    target_width: int = 220
    palette_colors: int = 64
    alpha_threshold: int = 20
    remove_bg_mode: str = "alpha"
    manual_bg_color: RGB = (200, 200, 200)
    bg_threshold: int = 22

def image_to_pixel_grid(img, opts: MosaicOptions) -> List[List[Pixel]]:
    im = img.convert("RGBA")
    if opts.remove_bg_mode in ("auto-corners", "manual-color"):
        if opts.remove_bg_mode == "auto-corners":
            w, h = im.size
            sample = [im.getpixel((x, y))[:3] for (x, y) in [(0,0), (w-1,0), (0,h-1), (w-1,h-1)]]
            ref = tuple(int(sum(c)/len(c)) for c in zip(*sample))
        else:
            ref = opts.manual_bg_color
        px = im.load()
        w, h = im.size
        thr = opts.bg_threshold ** 2
        for y in range(h):
            for x in range(w):
                r, g, b, a = px[x, y]
                if a > 0 and (r-ref[0])**2 + (g-ref[1])**2 + (b-ref[2])**2 <= thr:
                    px[x, y] = (r, g, b, 0)

    target_width = max(10, int(opts.target_width))
    target_height = max(10, int(target_width * im.height / im.width))
    small = im.resize((target_width, target_height), Image.Resampling.LANCZOS)

    white_bg = Image.new("RGBA", small.size, (255,255,255,255))
    comp = Image.alpha_composite(white_bg, small)
    quantized = comp.convert("RGB").convert("P", palette=Image.Palette.ADAPTIVE, colors=int(opts.palette_colors)).convert("RGB")

    grid: List[List[Pixel]] = []
    px_small = small.load()
    px_quant = quantized.load()
    for y in range(small.height):
        row: List[Pixel] = []
        for x in range(small.width):
            a = px_small[x, y][3]
            row.append(None if a <= opts.alpha_threshold else px_quant[x, y])
        grid.append(row)
    return grid

# ---------------- Crisp Pixel Banner ----------------
@dataclass
class BannerOptions:
    cols: int
    rows: int = 27
    bg_color: RGB = EXCEL_GREEN
    text_color: RGB = WHITE
    accent_color: RGB = WHITE
    grid_texture: bool = True
    style: str = "Pixel"  # only pixel renderer to guarantee clarity
    headline: str = "40 LAT EXCELA"
    subline: str = "Wszystkiego najlepszego! — Piotr"

def banner_to_pixel_grid(opts: BannerOptions) -> List[List[RGB]]:
    cols, rows = opts.cols, opts.rows
    bg = opts.bg_color
    grid: List[List[RGB]] = [[bg for _ in range(cols)] for _ in range(rows)]

    # Excel-like subtle grid texture
    if opts.grid_texture and rows >= 12 and cols >= 40:
        tex = lighten(bg, 0.15)
        step = 6
        for x in range(0, cols, step):
            for y in range(rows):
                grid[y][x] = tex
        for y in range(0, rows, step):
            for x in range(cols):
                grid[y][x] = tex

    # Decide scales to fit
    max_scale = max(1, rows // 7)          # how many cell-rows per glyph row
    scale_head = max(1, min(max_scale, int(rows * 0.75 // 7)))
    if scale_head < 2 and max_scale >= 2:
        scale_head = 2  # ensure legibility for typical 20-row banner

    spacing_head = max(1, scale_head) // 1
    head = normalize_ascii(opts.headline)
    # shrink until fits width
    while text_size_cells(head, scale_head, spacing_head) > cols - 4 and scale_head > 1:
        scale_head -= 1

    # Place headline centered horizontally, near top
    head_w = text_size_cells(head, scale_head, spacing_head)
    x0 = max(2, (cols - head_w) // 2)
    y0 = max(1, rows//6 - (7*scale_head)//2)
    blit_text(grid, x0, y0, head, opts.text_color, scale=scale_head, spacing=spacing_head)

    # Subline (smaller)
    sub = normalize_ascii(opts.subline)
    scale_sub = max(1, min(max_scale, int(rows * 0.35 // 7)))
    spacing_sub = 1
    # shrink to fit
    while text_size_cells(sub, scale_sub, spacing_sub) > cols - 4 and scale_sub > 1:
        scale_sub -= 1
    sub_w = text_size_cells(sub, scale_sub, spacing_sub)
    xs = max(2, (cols - sub_w) // 2)
    ys = rows - (7*scale_sub) - max(1, rows//10)
    blit_text(grid, xs, ys, sub, opts.text_color, scale=scale_sub, spacing=spacing_sub)

    return grid

# ---------------- Excel writer ----------------
@dataclass
class CellGeometry:
    col_width: float = 2.15
    row_height: float = 12.15
    margin_top: int = 2
    margin_left: int = 2
    spacer_rows: int = 2

class FillCache:
    def __init__(self):
        self.cache: Dict[str, PatternFill] = {}
    def get(self, rgb: RGB) -> PatternFill:
        key = to_hex_argb(rgb)
        if key not in self.cache:
            self.cache[key] = PatternFill(start_color=key, end_color=key, fill_type="solid")
        return self.cache[key]

def ensure_square_cells(ws, start_row: int, rows: int, start_col: int, cols: int, geom: CellGeometry):
    for c in range(start_col, start_col + cols):
        ws.column_dimensions[get_column_letter(c)].width = geom.col_width
    for r in range(start_row, start_row + rows):
        ws.row_dimensions[r].height = geom.row_height

def paint_pixels(ws, start_row: int, start_col: int, grid: List[List[Pixel]], fill_cache: FillCache, paint_background: Optional[RGB] = None):
    rows = len(grid)
    cols = len(grid[0]) if rows else 0
    for y in range(rows):
        for x in range(cols):
            rgb = grid[y][x]
            if rgb is None:
                if paint_background is None:
                    continue
                rgb = paint_background
            ws.cell(row=start_row + y, column=start_col + x).fill = fill_cache.get(rgb)

def build_workbook(portrait: List[List[Pixel]], banner: Optional[List[List[RGB]]], geom: CellGeometry, layout: str = "vertical", background: Optional[RGB] = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Excel Pixel Art"
    ws.sheet_view.showGridLines = False

    fill_cache = FillCache()
    start_row = geom.margin_top
    start_col = geom.margin_left

    banner_has_content = banner is not None and len(banner) > 0 and len(banner[0]) > 0

    if layout == "vertical":
        h1 = len(portrait); w1 = len(portrait[0]) if h1 else 0
        ensure_square_cells(ws, start_row, h1, start_col, w1, geom)
        paint_pixels(ws, start_row, start_col, portrait, fill_cache, paint_background=background)
        if banner_has_content:
            for r in range(start_row + h1, start_row + h1 + geom.spacer_rows):
                ws.row_dimensions[r].height = 6
            h2 = len(banner);   w2 = len(banner[0])   if h2 else 0
            ensure_square_cells(ws, start_row + h1 + geom.spacer_rows, h2, start_col, w2, geom)
            paint_pixels(ws, start_row + h1 + geom.spacer_rows, start_col, banner, fill_cache, paint_background=background)
    else:
        h1 = len(portrait); w1 = len(portrait[0]) if h1 else 0
        if banner_has_content:
            h2 = len(banner);   w2 = len(banner[0])   if h2 else 0
            rows = max(h1, h2)
            total_cols = w1 + geom.spacer_rows + w2
            ensure_square_cells(ws, start_row, rows, start_col, total_cols, geom)
            paint_pixels(ws, start_row, start_col, portrait, fill_cache, paint_background=background)
            paint_pixels(ws, start_row, start_col + w1 + geom.spacer_rows, banner, fill_cache, paint_background=background)
        else:
            ensure_square_cells(ws, start_row, h1, start_col, w1, geom)
            paint_pixels(ws, start_row, start_col, portrait, fill_cache, paint_background=background)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# ---------------- Streamlit UI ----------------
st.set_page_config(page_title="Excel Pixel Art — 40 lat Excela (crisp)", page_icon=":bar_chart:", layout="wide")
st.title("🎉 Excel Pixel Art - Świętujemy 40 lat z MS Excel")

with st.sidebar:
    st.header("⚙️ Ustawienia")
    target_width = st.slider("Szerokość (kolumny)", 60, 220, 220, 2)
    palette_colors = st.slider("Liczba kolorów (paleta)", 8, 64, 64, 2)
    alpha_threshold = st.slider("Próg przezroczystości (PNG alpha)", 0, 255, 20, 5)
    remove_bg_mode = st.selectbox("Usuwanie tła", ["alpha", "auto-corners", "manual-color"], index=0)
    manual_bg_hex = st.color_picker("Kolor tła (manual)", "#C8C8C8")
    bg_threshold = st.slider("Tolerancja usuwania tła", 0, 80, 22)

    st.markdown("---")
    st.subheader("Baner")
    banner_enabled = st.checkbox("Generuj baner", value=True)
    banner_rows = st.slider("Wysokość banera (wiersze)", 12, 40, 27, 1, disabled=not banner_enabled)
    banner_bg_hex = st.color_picker("Kolor tła banera", "#107C41", disabled=not banner_enabled)
    banner_text_color_hex = st.color_picker("Kolor tekstu", "#FFFFFF", disabled=not banner_enabled)
    banner_accent_hex = st.color_picker("Kolor akcentu", "#FFFFFF", disabled=not banner_enabled)  # kept for API compatibility
    grid_texture = st.checkbox("Tekstura siatki Excela", value=True, disabled=not banner_enabled)
    headline = st.text_input("Nagłówek", value="40 LAT EXCELA", disabled=not banner_enabled)
    subline = st.text_input("Podtytuł", value="Wszystkiego najlepszego! — Piotr", disabled=not banner_enabled)

    if not banner_enabled:
        st.caption("Baner wyłączony – wygenerujemy tylko mozaikę ze zdjęcia.")

    st.markdown("---")
    st.subheader("Arkusz")
    layout = st.selectbox("Układ", ["vertical (góra+dół)", "side-by-side (obok siebie)"], index=0)
    col_width = st.slider("Szerokość kolumn (Excel)", 1.2, 4.0, 2.15, 0.05)
    row_height = st.slider("Wysokość wierszy (pkt)", 8.0, 22.0, 12.15, 0.1)
    paint_background = st.checkbox("Maluj tło arkusza", value=False)
    bg_fill_hex = st.color_picker("Kolor tła arkusza", "#FFFFFF")
    spacer_rows = st.slider("Przerwa między portretem a banerem", 0, 6, 2)

uploaded = st.file_uploader("Wrzuć zdjęcie, ustaw parametry mozaiki, sprawdź podgląd i pobierz gotowe dzieło w MS Excel 😁", type=["png", "jpg", "jpeg", "webp"])

def hex_to_rgb(hx: str) -> RGB:
    hx = hx.strip()
    if hx.startswith("#"): hx = hx[1:]
    if len(hx) == 3: hx = "".join([c*2 for c in hx])
    return (int(hx[0:2], 16), int(hx[2:4], 16), int(hx[4:6], 16))

col_preview, col_actions = st.columns([2,1])

if uploaded is not None:
    img = Image.open(uploaded)
    col_preview.image(img, caption="Podgląd zdjęcia", use_container_width=True)

    mo = MosaicOptions(
        target_width=target_width,
        palette_colors=palette_colors,
        alpha_threshold=alpha_threshold,
        remove_bg_mode=remove_bg_mode,
        manual_bg_color=hex_to_rgb(manual_bg_hex),
        bg_threshold=bg_threshold
    )
    portrait_grid = image_to_pixel_grid(img, mo)

    banner_grid: Optional[List[List[RGB]]] = None
    if banner_enabled:
        bo = BannerOptions(
            cols=len(portrait_grid[0]),
            rows=banner_rows,
            bg_color=hex_to_rgb(banner_bg_hex),
            text_color=hex_to_rgb(banner_text_color_hex),
            accent_color=hex_to_rgb(banner_accent_hex),
            grid_texture=grid_texture,
            headline=headline if headline.strip() else "40 LAT EXCELA",
            subline=subline if subline.strip() else "Wszystkiego najlepszego! — Piotr",
        )
        banner_grid = banner_to_pixel_grid(bo)

    # Previews
    def grid_to_image(grid: List[List[Pixel]], bg: Optional[RGB]) -> Image.Image:
        h = len(grid); w = len(grid[0]) if h else 0
        im = Image.new("RGB", (w, h), bg if bg is not None else WHITE)
        draw = ImageDraw.Draw(im)
        for y in range(h):
            for x in range(w):
                pix = grid[y][x]
                if pix is None:
                    if bg is None: 
                        continue
                    pix = bg
                draw.point((x, y), fill=pix)
        return im

    preview_bg = hex_to_rgb(bg_fill_hex) if paint_background else None
    preview_portrait = grid_to_image(portrait_grid, preview_bg).resize((len(portrait_grid[0])*4, len(portrait_grid)*4), Image.NEAREST)
    preview_banner = None
    if banner_grid is not None:
        preview_banner = grid_to_image(banner_grid, preview_bg).resize((len(banner_grid[0])*4, len(banner_grid)*4), Image.NEAREST)

    with col_preview:
        st.subheader("Podgląd mozaiki (piksele)")
        st.image(preview_portrait, caption="Portret (siatka komórek)", use_container_width=True)
        if preview_banner is not None:
            st.image(preview_banner, caption='Baner „40 lat Excela” (pixel-perfect)', use_container_width=True)
        else:
            st.caption("Baner został pominięty w tym projekcie.")

    with col_actions:
        st.subheader("Generowanie pliku")
        geom = CellGeometry(col_width=col_width, row_height=row_height, margin_top=2, margin_left=2, spacer_rows=spacer_rows)
        layout_choice = "vertical" if layout.startswith("vertical") else "horizontal"
        bg_for_sheet = hex_to_rgb(bg_fill_hex) if paint_background else None

        total_cells = len(portrait_grid) * len(portrait_grid[0])
        if banner_grid is not None:
            total_cells += len(banner_grid) * len(banner_grid[0])
        st.caption(f"Szacunkowa liczba komórek: {total_cells:,}")

        if st.button("🧩 Generuj plik Excel (.xlsx)"):
            xlsx_bytes = build_workbook(portrait_grid, banner_grid, geom, layout=layout_choice, background=bg_for_sheet)
            st.success("Gotowe! Pobierz plik poniżej:")
            st.download_button("⬇️ Pobierz Excel", data=xlsx_bytes, file_name="Excel_40_lat_pixel_art.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
else:
    st.info("➡️ Wgraj obraz, aby zacząć. Najlepiej PNG z przezroczystością – wtedy tło w Excelu będzie czyste. ")
    st.markdown(
        """
        <div style='text-align: right; margin-top: 40px; font-size: 15px;'>
            Najlepsze życzenia z okazji 40-lecia Excela składa<br/>
            <strong>Piotr Szanser</strong>
            <a href="https://www.linkedin.com/in/pszanser/" target="_blank" style="text-decoration: none;">
                <img src="https://cdn.jsdelivr.net/gh/devicons/devicon/icons/linkedin/linkedin-original.svg" alt="LinkedIn" style="height: 20px; vertical-align: middle; margin-bottom: 2px;" />
            </a>
        </div>
        """,
        unsafe_allow_html=True,
        )
